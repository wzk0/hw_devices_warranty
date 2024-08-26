import ddddocr
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service
from json import dumps
from time import sleep
from requests import get
import openpyxl
import ctypes
import random
import os
import string
import shutil

ocr = ddddocr.DdddOcr()
# 序列号列表
base = openpyxl.load_workbook('base.xlsx')
# 选择活动的工作表
sheet = base.active

# 读取第一列的所有数据
ids = []
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
    for cell in row:
        ids.append(cell.value)
base.close()

# 输出
error_ids = []
success_ids = []


# 输出报错情况，并将出错序列号添加到error_ids
def error(reason, _id):
    print('%s Error - %s' % (reason, _id))
    error_ids.append(_id)


# 将error_ids与success_ids写入文件
def save():
    error_book = openpyxl.Workbook()
    sheet = error_book.active
    for index, value in enumerate(error_ids, start=1):
        sheet.cell(row=index, column=1, value=value)
    error_book.save('error.xlsx')
    error_book.close()
    with open('success.json', 'w', encoding='utf-8') as f:
        f.write(dumps(success_ids, ensure_ascii=False))
    if not os.path.exists('outputs'):
        os.makedirs('outputs')
    shutil.move('success.json', os.path.join('outputs', ''.join(random.choices(string.ascii_letters + string.digits, k=8)) + '.json'))
    ctypes.windll.user32.MessageBoxW(0,
                                     "查询完毕！错误序列号已写入 error.xlsx，请在重命名 error.xlsx 为 base.xlsx 后重新运行程序\n【此轮成功查询个数：%s 失败个数：%s】" % (
                                         len(success_ids), len(error_ids)), "提示", 0x40 | 0x1)


GECKODRIVER_PATH = r'./geckodriver.exe'
service = Service(executable_path=GECKODRIVER_PATH, service_args=['--marionette-port', '2828', '--connect-existing'])
browser = webdriver.Firefox(service=service)

for i in ids:
    # 无效序列号头，减少重复请求次数
    if i.startswith('LQ22B'):
        pass
    else:
        browser.get("https://consumer.huawei.com/cn/support/warranty-query/")
        # 休眠0.8s等待网页加载
        sleep(0.8)
        browser.find_element(By.CLASS_NAME, 'wic-device-wrap').find_element(By.TAG_NAME, 'input').send_keys(i)
        photo_url = browser.find_element(By.CLASS_NAME, 'warranty-inquiry').find_element(By.TAG_NAME,
                                                                                         'img').get_attribute('src')
        try:
            # 尝试写入验证码图片到本地，ocr获取结果
            with open('photo.jpg', 'wb') as p:
                p.write(get(photo_url).content)
            result = ocr.classification(open('photo.jpg', "rb").read())
        except:
            # 写入失败的处理，原因往往是网络不稳定导致验证码图片没有在0.8s内写入到本地，同时将验证码ocr结果赋值1
            result = '1'
        # 如果验证码长度不足4位则直接将序列号记录
        if len(result) != 4:
            error('CAPTCHA Code %s' % result, i)
        else:
            browser.find_element(By.CLASS_NAME, 'wic-identify').find_element(By.TAG_NAME, 'input').send_keys(result)
            browser.find_element(By.CLASS_NAME, 'warranty-inquiry-btn').find_element(By.TAG_NAME, 'a').click()
            # 点击查询按钮后休眠1s等待内容加载
            sleep(1)
            try:
                text = browser.find_element(By.CLASS_NAME, 'warranty-content').find_element(By.CLASS_NAME,
                                                                                            'warranty-content-guarantee').text.split(
                    '\n')[1:]
                success_ids.append({'设备序列号': i, '设备名称': text[0], '序列号': text[1].split('：')[-1],
                                    '预估保修截止日期': text[2].split('：')[-1], '可使用范围': text[3].split('：')[-1]})
                print('Successful ID - %s - %s' % (i, ids.index(i)))
            except:
                # 尝试将结果写入success_ids，失败原因1：网络不稳定导致未在1s内加载完毕；失败原因2：序列号无效
                error('Getting Text - %s' % result, i)

save()
